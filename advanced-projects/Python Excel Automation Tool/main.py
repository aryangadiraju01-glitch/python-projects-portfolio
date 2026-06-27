import openpyxl as xl
from openpyxl.chart import BarChart, Reference

name_of_file = ""
discount_percentage = 0

# Validate Excel file extension
is_valid = True

while is_valid:
    name_of_file = input("Enter the name of the file: ")

    if not name_of_file.endswith(".xlsx"):
        print("Invalid file extension. Please try again.")
    else:
        is_valid = False

# Validate discount percentage
percent_valid = True

while percent_valid:
    discount_percentage = float(input("Enter the discount %: "))

    if discount_percentage < 0 or discount_percentage > 100:
        print("Invalid number. Please enter a valid percentage.")
    else:
        percent_valid = False


def process_workbook(filename, discount):
    # Load workbook and select sheet
    wb = xl.load_workbook(filename)
    sheet = wb["Sheet1"]

    # Convert percentage discount into multiplier
    discount = 1 - (discount * 0.01)

    # Add header for discounted prices
    column_d = sheet.cell(1, 4)
    column_d.value = "Discounted Prices"

    # Calculate discounted prices
    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        corrected_price = round(cell.value * discount, 2)

        discounted_price_cell = sheet.cell(row, 4)
        discounted_price_cell.value = corrected_price

    # Create bar chart from discounted prices
    values = Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=4, max_col=4)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart)

    # Save updated workbook
    output_filename = "corrected_" + filename
    wb.save(output_filename)


process_workbook(name_of_file, discount_percentage)